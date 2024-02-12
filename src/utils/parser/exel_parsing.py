from loguru import logger
from openpyxl import load_workbook
from openpyxl.reader.excel import ExcelReader


class ParseExel:
    """
    Парсер exel-файла с данными о меню, подменю и блюдах
    """

    __DATA: dict = {}

    @classmethod
    def new_dict(cls):
        cls.__DATA = {}

    @classmethod
    def _open_file(cls, file: str) -> ExcelReader | bool:
        """
        Метод открывает и возвращает объект exel-файла
        :param file: директория для открытия файла
        :return: объект файла с данными
        """
        try:
            wb = load_workbook(file)  # Объект exel-файла
            sheet = wb['Меню']  # Объект нужной страницы

            return sheet

        except FileNotFoundError:
            logger.error(f'Файл для парсинга данных не найден: {file}')

            return False

    @classmethod
    def _check_row(cls, sheet) -> bool:
        """
        Метод проверяет наличие данных на обновление
        :param sheet: объект страницы с данными
        :return: bool
        """
        if sheet.max_row > 0:
            return True

        return False

    @classmethod
    def _parse_dishes(cls, sheet, row) -> list:
        """
        Метод для парсинга данных о блюдах
        :param sheet: объект страницы с данными
        :param row: номер строки для парсинга
        :return: список с блюдами
        """
        dishes_list = []

        for line in range(row, sheet.max_row + 1):
            num_dish = sheet[f'C{line}'].value

            if isinstance(num_dish, int):
                discount = sheet[f'G{line}'].value

                dishes_list.append(
                    {
                        'number': num_dish,
                        'title': sheet[f'D{line}'].value,
                        'description': sheet[f'E{line}'].value,
                        'price': sheet[f'F{line}'].value,
                        'discount': discount if discount else 0
                    }
                )

            else:
                break

        return dishes_list

    @classmethod
    def _parse_submenu(cls, sheet, row) -> list:
        """
        Метод для парсинга данных о подменю
        :param sheet: объект страницы с данными
        :param row: номер строки для парсинга
        :return: список с подменю
        """

        submenus_list = []

        for line in range(row + 1, sheet.max_row + 1):
            num_submenu = sheet[f'B{line}'].value

            if isinstance(num_submenu, int):
                dishes = cls._parse_dishes(sheet=sheet, row=line + 1)

                submenus_list.append(
                    {
                        'number': num_submenu,
                        'title': sheet[f'C{line}'].value,
                        'description': sheet[f'D{line}'].value,
                        'dishes': dishes
                    }
                )

            elif num_submenu is None:
                continue

            else:
                break

        return submenus_list

    @classmethod
    def _parse_menu(cls, sheet) -> None:
        """
        Метод для парсинга данных о меню
        :param sheet: объект страницы с данными
        :return: None
        """

        cls.__DATA['menus'] = []

        for row in range(2, sheet.max_row + 1):
            num_menu = sheet[f'A{row}'].value

            if isinstance(num_menu, int):
                submenus = cls._parse_submenu(sheet=sheet, row=row)

                cls.__DATA['menus'].append(
                    {
                        'number': num_menu,
                        'title': sheet[f'B{row}'].value,
                        'description': sheet[f'C{row}'].value,
                        'submenus': submenus
                    }
                )

    @classmethod
    def parse_data(cls, file: str) -> dict:
        """
        Метод для парсинга данных о меню с exel-файла
        :param file: файл для парсинга данных
        :return: словарь с данными
        """
        cls.new_dict()
        sheet = cls._open_file(file=file)

        if sheet and cls._check_row(sheet=sheet):
            cls._parse_menu(sheet=sheet)
        else:
            logger.warning('Нет данных для обновления меню')

        return cls.__DATA
